#import the  csv module
import csv
#import the load_workbook module from package openpyxl
from openpyxl import load_workbook

#creating the woorkbook object
wb=load_workbook(filename='C:/Users/oper/Downloads/book3.xlsx')
sheet=wb.active #active the woorkbook
csv_data=[] #creating list object to store values
for value in sheet.iter_rows(values_only=True): #values=True indicates the taking present values not empty
    csv_data.append(list(value))
    
with open('sampledata.csv','w')  as csv_obj: #creating a new file to save csv data in a file
    writer=csv.writer(csv_obj,delimiter=',')
    for line in csv_data:
        writer.writerow(line)
